import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook, Workbook

ws=load_workbook('files/exp_data.xlsx')['Sheet1']
max_row=ws.max_row

data = []
for i in range(1, max_row+1):
	data.append([ws[f"A{i}"].value, ws[f"B{i}"].value])

data=np.array(data)

x=data[:, 0]
y=data[:, 1]

dy=y[1:]-y[:-1]
dx=x[1:]-x[:-1]

ydot=y.copy()
ydot[1:]=dy/dx
ydot[0]=ydot[1]

d2y=dy[1:]-dy[:-1]

yddot=y.copy()
yddot[2:]=d2y/(dx[1:])**2
yddot[0]=yddot[1]

plt.plot(x,y, label="Exp")
plt.plot(x,ydot, label='d/dx(Exp)')
plt.plot(x,yddot, label='d^2/dx^2(Exp)')
plt.xlabel('x')
plt.ylabel('y')
plt.title('Exp')
plt.legend()
plt.show()

my_wb=Workbook()
my_ws=my_wb.active

j=0
for x_i in x:
	
	my_ws[f"A{i+1}"]=x_i
	my_ws[f"B{i+1}"]=yddot[j]

	j+=1

my_wb.save("files/dtwoexp.xlsx")